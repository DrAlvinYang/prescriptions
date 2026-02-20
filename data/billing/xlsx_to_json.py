#!/usr/bin/env python3
"""Convert billing_codes.xlsx and diagnostic_codes.xlsx to their JSON files.

Usage:
    python3 xlsx_to_json.py
    python3 xlsx_to_json.py --verbose

Reads:  data/billing_codes.xlsx    -> data/billing_codes.json
        data/diagnostic_codes.xlsx -> data/diagnostic_codes.json

Comma-separated values in array columns are split into arrays.
Empty cells become empty arrays (for arrays) or appropriate defaults.
"""

from __future__ import annotations

import json
import logging
import sys
from collections import Counter
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

logger = logging.getLogger(__name__)

SCRIPT_DIR = Path(__file__).resolve().parent


# -- Column maps --------------------------------------------------------------
# Column indices (A-O, 0-indexed) for each xlsx file.
# Group is derived from the sheet name, not a column.

BILLING_COL: dict[str, int] = {
    "subgroups": 0,
    "code": 1,
    "name": 2,
    "search_terms": 3,
    "fee": 4,
    "modifier_percentage": 5,
    "is_ortho_code": 6,
    "sedation_affiliated": 7,
    "has_c_code": 8,
    "sedation_base_units": 9,
    "related_modifiers": 10,
    "commonly_billed_with": 11,
    "conflicts_with": 12,
    "notes": 13,
    "hidden_notes": 14,
}

DIAG_COL: dict[str, int] = {
    "category": 0,
    "subcategory": 1,
    "code": 2,
    "name": 3,
    "search_terms": 4,
    "suggested_billing_codes": 5,
}


# -- Cell parsers --------------------------------------------------------------

def _is_blank(value: Any) -> bool:
    """True if value is None or whitespace-only."""
    return value is None or str(value).strip() == ""


def _cell(row: tuple, index: int) -> Any:
    """Safe column access -- returns None for out-of-bounds indices."""
    return row[index] if index < len(row) else None


def parse_str(value: Any) -> str:
    """Parse a string value, returning empty string if blank."""
    if _is_blank(value):
        return ""
    return str(value).strip()


def parse_bool(value: Any) -> bool:
    """Parse Yes/No string to boolean."""
    if _is_blank(value):
        return False
    return str(value).strip().lower() in ("yes", "true", "1")


def parse_fee(value: Any) -> float:
    """Parse fee as a float rounded to 2 decimal places, 0.0 if blank."""
    if _is_blank(value):
        return 0.0
    try:
        return round(float(value), 2)
    except (ValueError, TypeError):
        return 0.0


def parse_int_or_none(value: Any) -> int | None:
    """Parse an integer value, returning None if blank.

    Handles float-formatted cells (e.g., 6.0 -> 6) from Excel.
    """
    if _is_blank(value):
        return None
    try:
        return int(float(value))
    except (ValueError, TypeError):
        return None


def parse_delimited(value: Any, delimiter: str = ",") -> list[str]:
    """Split a delimited string into a cleaned list."""
    if _is_blank(value):
        return []
    return [item.strip() for item in str(value).split(delimiter) if item.strip()]


# -- Row parsers ---------------------------------------------------------------

def _extract_code(row: tuple, col_map: dict[str, int]) -> str | None:
    """Extract and validate code from a row. Returns None if row is empty/blank."""
    if not row:
        return None
    val = _cell(row, col_map["code"])
    if _is_blank(val):
        return None
    return str(val).strip()


def _parse_billing_row(row: tuple, group: str) -> dict | None:
    """Parse a single billing xlsx row into a code dict. Returns None if empty."""
    code = _extract_code(row, BILLING_COL)
    if not code:
        return None

    col = BILLING_COL
    entry = {
        "code": code,
        "name": parse_str(_cell(row, col["name"])),
        "fee": parse_fee(_cell(row, col["fee"])),
        "search_terms": parse_delimited(_cell(row, col["search_terms"]), delimiter=";"),
        "group": group,
        "subgroups": parse_delimited(_cell(row, col["subgroups"]), delimiter=";"),
        "related_modifiers": parse_delimited(_cell(row, col["related_modifiers"])),
        "commonly_billed_with": parse_delimited(_cell(row, col["commonly_billed_with"])),
        "conflicts_with": parse_delimited(_cell(row, col["conflicts_with"])),
        "notes": parse_str(_cell(row, col["notes"])),
        "hidden_notes": parse_str(_cell(row, col["hidden_notes"])),
        "is_ortho_code": parse_bool(_cell(row, col["is_ortho_code"])),
        "sedation_affiliated": parse_bool(_cell(row, col["sedation_affiliated"])),
        "sedation_base_units": parse_int_or_none(_cell(row, col["sedation_base_units"])),
        "has_c_code": parse_bool(_cell(row, col["has_c_code"])),
    }

    pct = parse_int_or_none(_cell(row, col["modifier_percentage"]))
    if pct is not None:
        entry["modifier_percentage"] = pct

    return entry


def _parse_diagnostic_row(row: tuple) -> dict | None:
    """Parse a single diagnostic xlsx row into a code dict. Returns None if empty."""
    code = _extract_code(row, DIAG_COL)
    if not code:
        return None

    col = DIAG_COL
    return {
        "code": code,
        "name": parse_str(_cell(row, col["name"])),
        "subcategory": parse_str(_cell(row, col["subcategory"])),
        "search_terms": parse_delimited(_cell(row, col["search_terms"]), delimiter=";"),
        "suggested_billing_codes": parse_delimited(_cell(row, col["suggested_billing_codes"])),
        "category": parse_str(_cell(row, col["category"])),
    }


# -- Core conversion ----------------------------------------------------------

def _write_json(data: list[dict], path: Path) -> None:
    """Sort by code and write JSON (does not mutate input)."""
    sorted_data = sorted(data, key=lambda c: c["code"])
    path.write_text(
        json.dumps(sorted_data, indent=2, ensure_ascii=False) + "\n",
        encoding="utf-8",
    )


# -- Public converters ---------------------------------------------------------

def convert_billing() -> bool:
    """Convert billing_codes.xlsx to billing_codes.json. Returns True on success."""
    xlsx_path = SCRIPT_DIR / "billing_codes.xlsx"
    json_path = SCRIPT_DIR / "billing_codes.json"

    try:
        wb = load_workbook(xlsx_path, read_only=True)
    except FileNotFoundError:
        logger.error("File not found: %s", xlsx_path)
        return False

    try:
        codes: list[dict] = []
        for ws in wb.worksheets:
            group = ws.title
            for row in ws.iter_rows(min_row=2, values_only=True):
                entry = _parse_billing_row(row, group)
                if entry:
                    codes.append(entry)
    finally:
        wb.close()

    _write_json(codes, json_path)

    groups = Counter(c["group"] for c in codes)
    filled_search = sum(1 for c in codes if c["search_terms"])
    logger.info("Billing: %d codes -> %s", len(codes), json_path.name)
    logger.info("  search_terms filled: %d/%d", filled_search, len(codes))
    for g, n in sorted(groups.items()):
        logger.info("  %s: %d codes", g, n)

    return True


def convert_diagnostic() -> bool:
    """Convert diagnostic_codes.xlsx to diagnostic_codes.json. Returns True on success."""
    xlsx_path = SCRIPT_DIR / "diagnostic_codes.xlsx"
    json_path = SCRIPT_DIR / "diagnostic_codes.json"

    try:
        wb = load_workbook(xlsx_path, read_only=True)
    except FileNotFoundError:
        logger.error("File not found: %s", xlsx_path)
        return False

    try:
        ws = wb.active
        codes: list[dict] = []
        # Row 1 = headers, Row 2 = description row, Row 3+ = data
        for row in ws.iter_rows(min_row=3, values_only=True):
            entry = _parse_diagnostic_row(row)
            if entry:
                codes.append(entry)
    finally:
        wb.close()

    _write_json(codes, json_path)

    filled_search = sum(1 for c in codes if c["search_terms"])
    filled_billing = sum(1 for c in codes if c["suggested_billing_codes"])
    logger.info("Diagnostic: %d codes -> %s", len(codes), json_path.name)
    logger.info("  search_terms filled: %d/%d", filled_search, len(codes))
    logger.info("  suggested_billing_codes filled: %d/%d", filled_billing, len(codes))

    return True


# -- CLI -----------------------------------------------------------------------

def main() -> int:
    """Convert all xlsx files to JSON. Returns 0 on success, 1 on failure."""
    logging.basicConfig(
        level=logging.INFO,
        format="%(levelname)s: %(message)s",
        handlers=[logging.StreamHandler(sys.stdout)],
    )

    logger.info("=" * 40)
    logger.info("xlsx_to_json")
    logger.info("=" * 40)

    success = convert_billing()
    if not convert_diagnostic():
        success = False

    if success:
        logger.info("Done.")
    else:
        logger.error("Completed with errors.")

    return 0 if success else 1


if __name__ == "__main__":
    sys.exit(main())
